import nflreadpy as nfl
import pandas as pd
import numpy as np
from sklearn.linear_model import Ridge
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

# --- CONFIGURATION ---
CURRENT_SEASON = 2025
CURRENT_WEEK = 18
STAGE_1_SEASONS = [2024, 2025]
QB_DECAY_RATE = 0.99
STAGE_2_SEASONS = [2025]
TEAM_DECAY_RATE = 0.95
MIN_STARTS_FOR_RATING = 3

# DYNAMIC THRESHOLDS
THRESH_KEY_NUM = 1.5  # If we cross 3 or 7
THRESH_STANDARD = 2.0  # If we are in "dead" zones

OUTPUT_FILE = f"NFL_Week_{CURRENT_WEEK}_KeyNumbers_Model.xlsx"

print(f"--- 🏈 NFL MODEL (Key Number Logic): Week {CURRENT_WEEK} ---")

# --- LOAD DATA ---
print("Fetching Data...")
try:
    schedule = nfl.load_schedules(seasons=STAGE_1_SEASONS).to_pandas()
    games = schedule[
        (schedule['game_type'].isin(['REG', 'POST'])) &
        (schedule['spread_line'].notnull()) &
        (schedule['result'].notnull())
        ].copy()

    current_slate = schedule[
        (schedule['season'] == CURRENT_SEASON) &
        (schedule['week'] == CURRENT_WEEK)
        ].copy()

    # Standardize to Negative = Favorite
    games['market_line'] = -games['spread_line']
    current_slate['market_line'] = -current_slate['spread_line']

    stats = nfl.load_player_stats(seasons=STAGE_1_SEASONS).to_pandas()
    if 'team' not in stats.columns: stats['team'] = stats['recent_team']
except Exception as e:
    print(f"Error: {e}")
    exit()

# --- MAP STARTERS ---
print("Mapping Starters...")


def get_starter_map(stats_df):
    passers = stats_df[stats_df['attempts'] > 0].copy()
    passers = passers.sort_values('attempts', ascending=False)
    starters = passers.groupby(['season', 'week', 'team']).head(1)
    starter_map = {}
    for _, row in starters.iterrows():
        starter_map[(row['season'], row['week'], row['team'])] = row['player_display_name']
    return starter_map


starter_lookup = get_starter_map(stats)


def assign_starters(row):
    h = starter_lookup.get((row['season'], row['week'], row['home_team']))
    a = starter_lookup.get((row['season'], row['week'], row['away_team']))
    return pd.Series([h, a])


games[['home_qb', 'away_qb']] = games.apply(assign_starters, axis=1)
games = games.dropna(subset=['home_qb', 'away_qb']).copy()

sorted_games = games.sort_values(['season', 'week'])
last_qb_map = {}
for _, row in sorted_games.iterrows():
    last_qb_map[row['home_team']] = row['home_qb']
    last_qb_map[row['away_team']] = row['away_qb']

# ==============================================================================
# CALCULATE RATINGS
# ==============================================================================
print("Calculating Ratings...")
games['s1_weight'] = games.apply(
    lambda x: QB_DECAY_RATE ** ((CURRENT_SEASON - x['season']) * 52 + (CURRENT_WEEK - x['week'])), axis=1)

h_team = pd.get_dummies(games['home_team'], dtype=int)
a_team = pd.get_dummies(games['away_team'], dtype=int)
h_qb = pd.get_dummies(games['home_qb'], dtype=int)
a_qb = pd.get_dummies(games['away_qb'], dtype=int)

all_teams = set(h_team.columns).union(a_team.columns)
all_qbs = set(h_qb.columns).union(a_qb.columns)

h_team = h_team.reindex(columns=all_teams, fill_value=0)
a_team = a_team.reindex(columns=all_teams, fill_value=0)
h_qb = h_qb.reindex(columns=all_qbs, fill_value=0)
a_qb = a_qb.reindex(columns=all_qbs, fill_value=0)

X = pd.concat([h_team.sub(a_team), h_qb.sub(a_qb)], axis=1)
X['HFA'] = 1
y = games['market_line']

clf_qbs = Ridge(alpha=1.5, fit_intercept=False)
clf_qbs.fit(X, y, sample_weight=games['s1_weight'])

coefs = pd.Series(clf_qbs.coef_, index=X.columns)
qb_ratings = coefs[list(all_qbs)]
qb_ratings = qb_ratings - qb_ratings.mean()

# GUARDRAILS
qb_counts = games['home_qb'].value_counts().add(games['away_qb'].value_counts(), fill_value=0)
bad_qb_rating = qb_ratings.quantile(0.95)  # High positive = Bad (in negative fav scale)

for qb in qb_ratings.index:
    if qb_counts.get(qb, 0) < MIN_STARTS_FOR_RATING:
        if qb_ratings[qb] < bad_qb_rating:
            qb_ratings[qb] = bad_qb_rating

qb_dict = qb_ratings.to_dict()

# TEAM RATINGS
games_s2 = games[games['season'] == CURRENT_SEASON].copy()


def remove_qb_impact(row):
    h_val = qb_dict.get(row['home_qb'], bad_qb_rating)
    a_val = qb_dict.get(row['away_qb'], bad_qb_rating)
    return row['market_line'] - (h_val - a_val)


games_s2['roster_line'] = games_s2.apply(remove_qb_impact, axis=1)
games_s2['s2_weight'] = games_s2.apply(lambda x: TEAM_DECAY_RATE ** (CURRENT_WEEK - x['week']), axis=1)

h_team_s2 = pd.get_dummies(games_s2['home_team'], dtype=int)
a_team_s2 = pd.get_dummies(games_s2['away_team'], dtype=int)
all_teams_s2 = set(h_team_s2.columns).union(a_team_s2.columns)
h_team_s2 = h_team_s2.reindex(columns=all_teams_s2, fill_value=0)
a_team_s2 = a_team_s2.reindex(columns=all_teams_s2, fill_value=0)

X_s2 = h_team_s2.sub(a_team_s2)
X_s2['HFA'] = 1
y_s2 = games_s2['roster_line']

clf_teams = Ridge(alpha=1.0, fit_intercept=False)
clf_teams.fit(X_s2, y_s2, sample_weight=games_s2['s2_weight'])

coefs_s2 = pd.Series(clf_teams.coef_, index=X_s2.columns)
hfa_final = coefs_s2['HFA']
team_ratings = coefs_s2.drop('HFA')

# ==============================================================================
# EXPORT
# ==============================================================================
print(f"Exporting...")

df_qbs = pd.DataFrame({'QB': qb_ratings.index, 'Rating': qb_ratings.values}).sort_values('QB', ascending=True)
df_teams = pd.DataFrame({'Team': team_ratings.index, 'Rating': team_ratings.values}).sort_values('Team', ascending=True)
backup_row = pd.DataFrame({'QB': ['(Generic Backup)'], 'Rating': [bad_qb_rating]})
df_qbs = pd.concat([backup_row, df_qbs], ignore_index=True)

with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    df_teams.to_excel(writer, sheet_name='DB_Teams', index=False)
    df_qbs.to_excel(writer, sheet_name='DB_QBs', index=False)
    # Store settings for Excel reference
    pd.DataFrame({'Metric': ['HFA', 'Thresh_Key', 'Thresh_Std'],
                  'Value': [hfa_final, THRESH_KEY_NUM, THRESH_STANDARD]}).to_excel(writer, sheet_name='Settings',
                                                                                   index=False)
    writer.book.create_sheet('Betting Board')

# ==============================================================================
# DASHBOARD
# ==============================================================================
wb = load_workbook(OUTPUT_FILE)
ws = wb['Betting Board']

# HEADERS
headers = [
    "Away Team", "Away QB", "Home Team", "Home QB",
    "Model (HOME)", "Vegas (HOME)", "Edge", "Req. Edge", "SIGNAL"
]
ws.append(headers)

# Styling
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
center = Alignment(horizontal="center", vertical="center")
input_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center

qb_dv = DataValidation(type="list", formula1="='DB_QBs'!$A$2:$A$300", allow_blank=True)
ws.add_data_validation(qb_dv)

row_idx = 2
for _, game in current_slate.iterrows():
    away_tm = game['away_team']
    home_tm = game['home_team']

    vegas_val = game['market_line']
    if pd.isna(vegas_val): vegas_val = "N/A"

    away_qb = last_qb_map.get(away_tm, "(Generic Backup)")
    home_qb = last_qb_map.get(home_tm, "(Generic Backup)")

    ws.cell(row=row_idx, column=1, value=away_tm).alignment = center
    ws.cell(row=row_idx, column=2, value=away_qb)
    ws.cell(row=row_idx, column=3, value=home_tm).alignment = center
    ws.cell(row=row_idx, column=4, value=home_qb)

    # Input Cell
    vegas_cell = ws.cell(row=row_idx, column=6, value=vegas_val)
    vegas_cell.fill = input_fill
    vegas_cell.alignment = center
    if vegas_val != "N/A": vegas_cell.number_format = '0.0'

    qb_dv.add(ws.cell(row=row_idx, column=2))
    qb_dv.add(ws.cell(row=row_idx, column=4))

    # Hidden Ratings
    ws.cell(row=row_idx, column=26, value=f"=VLOOKUP(A{row_idx},'DB_Teams'!A:B,2,FALSE)")
    ws.cell(row=row_idx, column=27, value=f"=VLOOKUP(B{row_idx},'DB_QBs'!A:B,2,FALSE)")
    ws.cell(row=row_idx, column=28, value=f"=VLOOKUP(C{row_idx},'DB_Teams'!A:B,2,FALSE)")
    ws.cell(row=row_idx, column=29, value=f"=VLOOKUP(D{row_idx},'DB_QBs'!A:B,2,FALSE)")

    # MODEL: (Home + HFA) - Away
    # Home is Negative (Good). Away is Positive (Bad).
    model_formula = f"=(AB{row_idx}+AC{row_idx}+Settings!$B$2) - (Z{row_idx}+AA{row_idx})"
    ws.cell(row=row_idx, column=5, value=model_formula).number_format = '0.0'
    ws.cell(row=row_idx, column=5).alignment = center
    ws.cell(row=row_idx, column=5).font = Font(bold=True)

    # EDGE: Model - Vegas
    if vegas_val != "N/A":
        ws.cell(row=row_idx, column=7, value=f"=E{row_idx}-F{row_idx}").number_format = '0.0'
        ws.cell(row=row_idx, column=7).alignment = center
    else:
        ws.cell(row=row_idx, column=7, value="--")

    # KEY NUMBER LOGIC (New Column H)
    # Check if we crossed 3, 7, -3, or -7
    # Logic: If (Model - 3) * (Vegas - 3) < 0, it means they are on opposite sides of 3.
    # Ref: E=Model, F=Vegas.
    t_key = "Settings!$B$3"  # 1.5
    t_std = "Settings!$B$4"  # 2.0

    kn_check = (
        f"OR("
        f"(E{row_idx}-3)*(F{row_idx}-3)<0, "  # Crossed 3
        f"(E{row_idx}-7)*(F{row_idx}-7)<0, "  # Crossed 7
        f"(E{row_idx}+3)*(F{row_idx}+3)<0, "  # Crossed -3
        f"(E{row_idx}+7)*(F{row_idx}+7)<0"  # Crossed -7
        f")"
    )

    req_edge_formula = f'=IF({kn_check}, {t_key}, {t_std})'
    ws.cell(row=row_idx, column=8, value=req_edge_formula).number_format = '0.0'
    ws.cell(row=row_idx, column=8).alignment = center

    # RECOMMENDATION (Column I)
    # Compare ABS(Edge) to Req Edge
    # If Edge > 0 (Positive), we like Away. If Edge > Req, Bet Away.
    # If Edge < 0 (Negative), we like Home. If Edge < -Req, Bet Home.
    rec_formula = (
        f'=IF(ISBLANK(F{row_idx}), "Enter Line", '
        f'IF(G{row_idx} < -H{row_idx}, "BET " & C{row_idx}, '  # Bet Home
        f'IF(G{row_idx} > H{row_idx}, "BET " & A{row_idx}, '  # Bet Away
        f'"PASS")))'
    )
    ws.cell(row=row_idx, column=9, value=rec_formula).font = Font(bold=True)
    ws.cell(row=row_idx, column=9).alignment = center

    row_idx += 1

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_font = Font(color="006100", bold=True)
ws.conditional_formatting.add(
    f"I2:I{row_idx}",
    CellIsRule(operator="beginsWith", formula=['"BET"'], stopIfTrue=True, fill=green_fill, font=green_font)
)

wb['DB_Teams'].sheet_state = 'hidden'
wb['DB_QBs'].sheet_state = 'hidden'
wb['Settings'].sheet_state = 'hidden'

wb.save(OUTPUT_FILE)
print(f"✅ Success! Generated Key Number Model in {OUTPUT_FILE}")
print("   > Column H 'Req. Edge' will show 1.5 if you cross a key number, 2.0 otherwise.")